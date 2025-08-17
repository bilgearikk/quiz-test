import asyncio
import json
import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from openpyxl import load_workbook

logger = logging.getLogger("quiz")
logging.basicConfig(level=logging.INFO)

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

class Player:
    def __init__(self, name: str, ws: WebSocket):
        self.name = name
        self.ws = ws
        self.score = 0
        self.answered_for_q: Dict[int, bool] = {}

class QuizState:
    def __init__(self):
        self.players: Dict[str, Player] = {}
        self.admins: Set[WebSocket] = set()
        self.questions: List[dict] = []
        self.current_q_index: int = -1
        self.accepting: bool = False
        self.q_started_at: Optional[datetime] = None
        self.q_duration_sec: int = 10
        self.round_active: bool = False

    def soft_reset(self):
        for p in self.players.values():
            p.score = 0
            p.answered_for_q.clear()
        self.current_q_index = -1
        self.accepting = False
        self.round_active = False
        self.q_started_at = None

STATE = QuizState()

def utc_now() -> datetime:
    return datetime.now(timezone.utc)

def load_questions_from_excel(path: str) -> List[dict]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "").strip().lower() for h in rows[0]]
    required = ["question", "option1", "option2", "option3", "option4", "correct_index"]
    hidx = {h: headers.index(h) for h in required}
    questions: List[dict] = []
    for r in rows[1:]:
        try:
            q_text = str(r[hidx["question"]] or "").strip()
            opts = [str(r[hidx[f"option{i+1}"]] or "").strip() for i in range(4)]
            cidx = max(0, min(3, int(r[hidx["correct_index"]] or 0)))
            if q_text and all(opts):
                questions.append({"question": q_text, "options": opts, "correct": cidx})
        except Exception as e:
            logger.warning("Excel satırı atlandı: %s", e)
    if not questions:
        raise ValueError("Excel'den geçerli soru bulunamadı.")
    return questions

def score_for_elapsed(elapsed: float) -> int:
    if elapsed <= 3.0:
        return 5
    elif elapsed <= 5.0:
        return 3
    elif elapsed <= 10.0:
        return 2
    return 0

async def broadcast(payload: dict):
    text = json.dumps(payload)
    dead_pids = []
    for pid, p in STATE.players.items():
        try:
            await p.ws.send_text(text)
        except:
            dead_pids.append(pid)
    for pid in dead_pids:
        STATE.players.pop(pid, None)

    dead_admins = []
    for a in STATE.admins:
        try:
            await a.send_text(text)
        except:
            dead_admins.append(a)
    for a in dead_admins:
        STATE.admins.discard(a)

async def start_question(index: int):
    STATE.current_q_index = index
    STATE.accepting = True
    STATE.round_active = True
    STATE.q_started_at = utc_now()
    q = STATE.questions[index]
    expires_at = STATE.q_started_at.timestamp() + STATE.q_duration_sec

    await broadcast({
        "type": "question",
        "index": index,
        "q_total": len(STATE.questions),
        "question": q["question"],
        "options": q["options"],
        "expires_at": expires_at,
    })

    asyncio.create_task(end_question_after_delay(STATE.q_duration_sec))

async def end_question_after_delay(delay: int):
    await asyncio.sleep(delay)
    await end_current_question()

async def end_current_question():
    if not STATE.round_active:
        return
    STATE.accepting = False
    STATE.round_active = False

    q = STATE.questions[STATE.current_q_index]
    await broadcast({"type": "reveal", "index": STATE.current_q_index, "correct": q["correct"]})
    await asyncio.sleep(2)

    if STATE.current_q_index + 1 < len(STATE.questions):
        await start_question(STATE.current_q_index + 1)
    else:
        leaderboard = sorted(
            ((p.name, p.score) for p in STATE.players.values()),
            key=lambda x: (-x[1], x[0].lower())
        )
        await broadcast({"type": "leaderboard", "all": list(leaderboard)})

async def check_all_answered():
    if not STATE.accepting or STATE.current_q_index < 0:
        return
    if all(p.answered_for_q.get(STATE.current_q_index) for p in STATE.players.values()):
        await end_current_question()

@app.get("/", response_class=HTMLResponse)
async def player_page(request: Request):
    return templates.TemplateResponse("player.html", {"request": request})

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    return templates.TemplateResponse("admin.html", {"request": request})

@app.get("/api/health")
async def health():
    return JSONResponse({"ok": True})

@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    pid = hex(id(ws))
    try:
        while True:
            raw = await ws.receive_text()
            try:
                msg = json.loads(raw)
            except:
                continue

            mtype = msg.get("type")

            if mtype == "join":
                name = (msg.get("name") or f"Player-{pid[-4:]}").strip()[:24]
                STATE.players[pid] = Player(name=name, ws=ws)
                await broadcast({"type": "lobby", "players": [p.name for p in STATE.players.values()]})
                await ws.send_text(json.dumps({"type": "joined", "name": name}))

            elif mtype == "admin":
                STATE.admins.add(ws)
                await ws.send_text(json.dumps({
                    "type": "admin_ack",
                    "players": [p.name for p in STATE.players.values()],
                    "q_count": len(STATE.questions),
                }))

            elif mtype == "load_questions":
                path = msg.get("path", "questions.xlsx")
                try:
                    STATE.questions = load_questions_from_excel(path)
                    await broadcast({"type": "questions_loaded", "count": len(STATE.questions)})
                except Exception as e:
                    await ws.send_text(json.dumps({"type": "error", "message": str(e)}))

            elif mtype == "start_quiz":
                if not STATE.questions:
                    await ws.send_text(json.dumps({"type": "error", "message": "Önce Excel'den soruları yükleyin."}))
                else:
                    STATE.soft_reset()
                    await start_question(0)

            elif mtype == "answer":
                if not STATE.accepting or STATE.current_q_index < 0:
                    continue
                player = STATE.players.get(pid)
                if not player or player.answered_for_q.get(STATE.current_q_index):
                    continue

                try:
                    chosen = int(msg.get("choice", -1))
                except:
                    chosen = -1

                q = STATE.questions[STATE.current_q_index]
                elapsed = (utc_now() - STATE.q_started_at).total_seconds() if STATE.q_started_at else 999
                if chosen == q["correct"] and elapsed <= STATE.q_duration_sec:
                    player.score += score_for_elapsed(elapsed)
                player.answered_for_q[STATE.current_q_index] = True

                await ws.send_text(json.dumps({
                    "type": "answer_ack",
                    "correct": chosen == q["correct"],
                    "score": player.score,
                }))

                await check_all_answered()

            elif mtype == "next":
                await end_current_question()

            elif mtype == "reset":
                STATE.soft_reset()
                await broadcast({"type": "reset_done"})

    except WebSocketDisconnect:
        STATE.players.pop(pid, None)
        STATE.admins.discard(ws)
        await broadcast({"type": "lobby", "players": [p.name for p in STATE.players.values()]})

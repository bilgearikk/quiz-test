# app.py
import asyncio
import json
import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set, Tuple
from io import BytesIO

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

logger = logging.getLogger("quiz")
logging.basicConfig(level=logging.INFO)

app = FastAPI()

# Static & templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# ---------------------- Game State ----------------------
class Player:
    def __init__(self, name: str, ws: WebSocket):
        self.name = name
        self.ws = ws
        self.score: int = 0
        self.answered_for_q: Dict[int, bool] = {}
        self.streak: int = 0  # art arda doğru sayısı


class QuizState:
    def __init__(self):
        self.players: Dict[str, Player] = {}   # key = connection id (hex)
        self.admins: Set[WebSocket] = set()
        self.questions: List[dict] = []
        self.current_q_index: int = -1
        self.accepting: bool = False
        self.q_started_at: Optional[datetime] = None
        self.q_duration_sec: int = 10
        self.round_active: bool = False

    def soft_reset(self):
        for p in self.players.values():
            p.score = 0
            p.answered_for_q.clear()
            p.streak = 0
        self.current_q_index = -1
        self.accepting = False
        self.round_active = False
        self.q_started_at = None


STATE = QuizState()


# ---------------------- Helpers ----------------------
def utc_now():
    return datetime.now(timezone.utc)


def _parse_rows(rows: List[List[object]]) -> List[dict]:
    if not rows:
        raise ValueError("Excel boş görünüyor.")
    headers = [str(h or "").strip().lower() for h in rows[0]]
    required = ["question", "option1", "option2", "option3", "option4", "correct_index"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise ValueError(f"Excel başlıkları eksik. Gerekli: {required}")
    idx = {h: headers.index(h) for h in required}

    out: List[dict] = []
    for r in rows[1:]:
        if not r or all(v is None for v in r):
            continue
        q_text = str(r[idx["question"]] or "").strip()
        options = [
            str(r[idx["option1"]] or "").strip(),
            str(r[idx["option2"]] or "").strip(),
            str(r[idx["option3"]] or "").strip(),
            str(r[idx["option4"]] or "").strip(),
        ]
        c_raw = r[idx["correct_index"]]
        correct = int(c_raw) if c_raw is not None else 0
        correct = max(0, min(3, correct))
        if not q_text or any(o == "" for o in options):
            continue
        out.append({"question": q_text, "options": options, "correct": correct})

    if not out:
        raise ValueError("Excel'den geçerli soru bulunamadı.")
    return out


def load_questions_from_excel(path: str) -> List[dict]:
    # Dosya yolundan oku
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return _parse_rows(rows)


def load_questions_from_bytes(data: bytes) -> List[dict]:
    # Upload edilen byte verisinden oku
    wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return _parse_rows(rows)


async def broadcast(payload: dict):
    """Send to all players and admins. Always JSON-encoded."""
    text = json.dumps(payload)
    dead_players: List[str] = []
    for pid, p in list(STATE.players.items()):
        try:
            await p.ws.send_text(text)
        except Exception:
            dead_players.append(pid)
    for pid in dead_players:
        STATE.players.pop(pid, None)

    dead_admins: List[WebSocket] = []
    for a in list(STATE.admins):
        try:
            await a.send_text(text)
        except Exception:
            dead_admins.append(a)
    for a in dead_admins:
        STATE.admins.discard(a)


async def broadcast_scores(topn: int = 5):
    leaderboard: List[Tuple[str, int]] = sorted(
        ((p.name, p.score) for p in STATE.players.values()),
        key=lambda x: x[1],
        reverse=True,
    )
    await broadcast({"type": "scores", "top5": list(leaderboard)[:topn]})


def score_for_elapsed(elapsed: float) -> int:
    # 0–3 sn => 5, 3–5 sn => 3, 5–10 sn => 2, aksi 0
    if elapsed <= 3.0:
        return 5
    elif elapsed <= 5.0:
        return 3
    elif elapsed <= 10.0:
        return 2
    return 0


async def start_question(index: int):
    STATE.current_q_index = index
    STATE.accepting = True
    STATE.round_active = True
    STATE.q_started_at = utc_now()

    q = STATE.questions[index]
    expires_at = (STATE.q_started_at.timestamp() + STATE.q_duration_sec)

    await broadcast({
        "type": "question",
        "index": index,
        "question": q["question"],
        "options": q["options"],
        "expires_at": expires_at,  # epoch seconds UTC
        "q_total": len(STATE.questions),
    })

    # Schedule to end the question after duration
    asyncio.create_task(end_question_after_delay(STATE.q_duration_sec))


async def end_question_after_delay(delay: int):
    await asyncio.sleep(delay)
    await end_current_question()


async def end_current_question():
    if not STATE.round_active:
        return
    STATE.accepting = False
    STATE.round_active = False

    # Reveal correct answer to everyone
    q = STATE.questions[STATE.current_q_index]
    await broadcast({
        "type": "reveal",
        "index": STATE.current_q_index,
        "correct": q["correct"],
    })

    # Short pause before next question
    await asyncio.sleep(2)

    # Next or leaderboard
    if STATE.current_q_index + 1 < len(STATE.questions):
        await start_question(STATE.current_q_index + 1)
    else:
        leaderboard = sorted(
            [(p.name, p.score) for p in STATE.players.values()],
            key=lambda x: x[1], reverse=True
        )
        top3 = leaderboard[:3]
        await broadcast({
            "type": "leaderboard",
            "top3": top3,
            "all": leaderboard,
        })


# ---------------------- HTTP Pages ----------------------
@app.get("/", response_class=HTMLResponse)
async def player_page(request: Request):
    return templates.TemplateResponse("player.html", {"request": request})


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    return templates.TemplateResponse("admin.html", {"request": request})


@app.get("/api/health")
async def health():
    return JSONResponse({"ok": True})


# ---------------------- Upload API ----------------------
@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    try:
        data = await file.read()
        STATE.questions = load_questions_from_bytes(data)
        await broadcast({"type": "questions_loaded", "count": len(STATE.questions)})
        return {"ok": True, "count": len(STATE.questions)}
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})


# ---------------------- WebSocket ----------------------
@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    pid = hex(id(ws))
    try:
        while True:
            raw = await ws.receive_text()
            try:
                msg = json.loads(raw)
            except json.JSONDecodeError:
                # Client sent non-JSON — ignore gracefully
                continue

            mtype = msg.get("type")

            # ---- Join as player ----
            if mtype == "join":
                name = (msg.get("name") or f"Player-{pid[-4:]}").strip()[:24]
                STATE.players[pid] = Player(name=name, ws=ws)
                # Notify admins/players about lobby change
                await broadcast({"type": "lobby", "players": [p.name for p in STATE.players.values()]})
                await ws.send_text(json.dumps({"type": "joined", "name": name}))
                # skor panelini güncelle
                await broadcast_scores()

            # ---- Join as admin ----
            elif mtype == "admin":
                STATE.admins.add(ws)
                await ws.send_text(json.dumps({
                    "type": "admin_ack",
                    "players": [p.name for p in STATE.players.values()],
                    "q_count": len(STATE.questions),
                }))

            # ---- Load questions from Excel path (opsiyonel) ----
            elif mtype == "load_questions":
                path = msg.get("path", "questions.xlsx")
                try:
                    STATE.questions = load_questions_from_excel(path)
                    await broadcast({"type": "questions_loaded", "count": len(STATE.questions)})
                except Exception as e:
                    await ws.send_text(json.dumps({"type": "error", "message": str(e)}))

            # ---- Start quiz (admin) ----
            elif mtype == "start_quiz":
                if not STATE.questions:
                    await ws.send_text(json.dumps({"type": "error", "message": "Önce Excel'den soruları yükleyin."}))
                else:
                    STATE.soft_reset()
                    await broadcast_scores()  # mini-leaderboard ilk gönderim
                    await start_question(0)

            # ---- Player answer ----
            elif mtype == "answer":
                if not STATE.accepting or STATE.current_q_index < 0:
                    continue
                player = STATE.players.get(pid)
                if not player:
                    continue

                # Only first answer per question
                if player.answered_for_q.get(STATE.current_q_index):
                    continue

                try:
                    chosen = int(msg.get("choice", -1))
                except (TypeError, ValueError):
                    chosen = -1

                q = STATE.questions[STATE.current_q_index]
                elapsed = (utc_now() - STATE.q_started_at).total_seconds() if STATE.q_started_at else 999

                was_correct = (chosen == q["correct"] and elapsed <= STATE.q_duration_sec)
                if was_correct:
                    player.score += score_for_elapsed(elapsed)
                    player.streak += 1
                else:
                    player.streak = 0

                player.answered_for_q[STATE.current_q_index] = True

                # feedback only to that player (UI anında yazı göstermiyor)
                await ws.send_text(json.dumps({
                    "type": "answer_ack",
                    "correct": was_correct,
                    "score": player.score,
                    "streak": player.streak,
                }))

                # canlı mini-leaderboard
                await broadcast_scores()

            # ---- Force end / next (admin) ----
            elif mtype == "next":
                await end_current_question()

            # ---- Reset lobby (admin) ----
            elif mtype == "reset":
                STATE.soft_reset()
                await broadcast({"type": "reset_done"})
                await broadcast_scores()

    except WebSocketDisconnect:
        # remove from players or admins
        STATE.players.pop(pid, None)
        STATE.admins.discard(ws)
        # notify lobby & mini scores update
        await broadcast({"type": "lobby", "players": [p.name for p in STATE.players.values()]})
        await broadcast_scores()
    except Exception as e:
        logger.exception("websocket error: %s", e)
